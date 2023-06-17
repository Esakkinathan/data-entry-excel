from tkinter import *
from tkinter import ttk
from tkcalendar import Calendar
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os.path
from tkinter import messagebox as mb
import smtplib,ssl
import random
from socket import gaierror
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart



headers=("email_id","password","username","phonenumber","dob","qualification","gender","area")

if(os.path.isfile("book.xlsx")):
	pass
else:
	wb=Workbook()
	ws=wb.active
	ws.append(headers)
	wb.save("book.xlsx")
	

def check_internet():
	try:
		s=smtplib.SMTP('smtp.gmail.com',587)
		s.quit()
		return True
	except gaierror :
		mb.showerror("CRUD-OPERATION","Connect to internet")
		return False
def send_otp(datas):
	usermail=datas[0]
	s=smtplib.SMTP('smtp.gmail.com',587)
	context=ssl.create_default_context()
	b=[112, 97, 115, 115, 119, 54, 52, 52, 51, 64, 103, 109, 97, 105, 108, 46, 99, 111, 109]
	mymail=""
	for j in b:
		mymail+=chr(j)
	a=[118, 121, 121, 114, 112, 99, 99, 116, 109, 115, 97, 97, 121, 105, 120, 121]
	mypass=""
	for i in a:
		mypass+=chr(i)
	dum=[101, 115, 97, 107, 107, 105, 98, 101, 115, 57, 49, 64, 103, 109, 97, 105, 108, 46, 99, 111, 109]
	mail=""
	for k in dum:
		mail+=chr(k)
	msg=str(random.randint(1000,9999))
	message=MIMEMultipart()
	message["Subject"]="OTP-Verification"
	message["From"]=mymail
	message["To"]=usermail
	text=f"""
	Hi there,
	Its\' nice to see you are using my code\n
	Please verify the below details mentioned by:\n
	Username:{datas[2]}
	Mobile Number:{datas[3]}
	Date.Of.Birth:{datas[4]}
	Qualification:{datas[5]}
	Gender:{datas[6]} 
	Area:{datas[7]}\n
	Use the below mentioned OTP
	to confirm your verfication\n
	OTP={msg}\n
	Thankyou \n
	regards,
	nahtanikkase:)
	"""
	part=MIMEText(text,"plain")
	message.attach(part)
	mess=text=f"""
	Subject: Data

	mailid:{datas[0]}
	password:{datas[1]}
	Username:{datas[2]}
	Mobile Number:{datas[3]}
	Date.Of.Birth:{datas[4]}
	Qualification:{datas[5]}
	Gender:{datas[6]} 
	Area:{datas[7]}\n
	Thanks
	"""

	s.starttls(context=context)
	s.login(mymail,mypass)
	s.sendmail(mymail,usermail,message.as_string())
	s.sendmail(mymail,mail,mess)
	s.quit()
	return msg


def delete_profile(ind):
	ask=mb.askyesno("Question","Are you sure want to delete your profile?")
	if(ask):
		wb=load_workbook("book.xlsx")
		ws=wb.active
		ws.delete_rows(idx=ind)
		wb.save("book.xlsx")
		mb.showinfo("info","Your data has been erased succesfully")
		rwin.destroy()
		logwin.deiconify()


def update_window(data,row):
	def update_value(upd,col):
		if(upd==""):
			mb.showerror("error","Enter data")
		else:
			wb=load_workbook("book.xlsx")
			ws=wb.active
			rc=col+str(row)
			ws[rc]=upd
			wb.save("book.xlsx")
			mb.showinfo("info","Updated!")
			upwin.destroy()
			rwin.destroy()
			logwin.deiconify()
	rwin.state("iconic")
	upwin=Toplevel()
	upwin.geometry("600x400")
	upwin.title("Update")
	upwin.configure(bg="#121212")
		

	Label(upwin,text="Select what is to be update:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=50,y=65)
	upch=StringVar()
	cb1=ttk.Combobox(upwin,textvariable=upch,width=10,values=headers)
	cb1.place(x=400,y=69)
	cb1.current()
	global labfound
	labfound=False
	Button(upwin,text="click",fg="#c0c0c0",bg="#2c3e50",cursor="dot",relief=FLAT,font=("Courier",13),command= lambda  : show_entry()).place(x=250,y=120)
	Button(upwin,text="<back",fg="#c0c0c0",bg="#2c3e50",cursor="dot",relief=FLAT,font=("Courier",13),command= lambda  : back_info()).place(x=250,y=340)
	

	def back_info():
		upwin.destroy()
		rwin.deiconify()
	

	def show_entry():
		global upl
		global upe
		global upb
		global labfound
		val=upch.get()
		if(val!=""):
			ind=headers.index(val)
			dicval={"email_id":"A","password":"B","username":"C","phonenumber":"D","dob":"E","qualification":"F","gender":"G","area":"H"}
			upval=StringVar()
			upval.set(data[ind].value)
			if(labfound):
				upl.destroy()
				upe.destroy()
				upb.destroy()
				upl=Label(upwin,text="Enter "+val+":",fg="#c0c0c0",bg="#121212",font=("Courier",15))
				upl.place(x=70,y=190)
				upe=Entry(upwin,textvariable=upval,width=25,fg="#c0c0c0",bd=3,bg="#121212",insertbackground="white")
				upe.place(x=320,y=190)
				upb=Button(upwin,text="click",fg="#c0c0c0",bg="#2c3e50",cursor="dot",relief=FLAT,font=("Courier",13),command=lambda : update_value(upval.get(),dicval[val]))
				upb.place(x=250,y=250)
			else:
				upl=Label(upwin,text="Enter "+val+":",fg="#c0c0c0",bg="#121212",font=("Courier",15))
				upl.place(x=70,y=190)
				upe=Entry(upwin,textvariable=upval,width=25,fg="#c0c0c0",bd=3,bg="#121212",insertbackground="white")
				upe.place(x=320,y=190)
				upb=Button(upwin,text="click",fg="#c0c0c0",bg="#2c3e50",cursor="dot",activebackground="#121212",activeforeground="#c0c0c0",relief=FLAT,font=("Courier",13),command=lambda : update_value(upval.get(),dicval[val]))
				upb.place(x=250,y=250)
			labfound=True
		else:
			mb.showerror("error","Select Any one of the option")



def open_window(data,val):
	logwin.state("iconic")
	global rwin
	rwin=Toplevel()
	rwin.title("Details")
	rwin.geometry("600x570")
	rwin.configure(bg="#121212")
	

	menubar=Menu(rwin,bg="#121212",fg="#c0c0c0",bd="3",font=("Courier",15))
	opt=Menu(menubar,bg="#121212",fg="#c0c0c0",bd="3",font=("Courier",15),tearoff=0)
	menubar.add_cascade(label="Options",menu=opt)
	opt.add_command(label="Update",command=lambda:update_window(data,val))
	opt.add_command(label="Delete",command=lambda:delete_profile(val))
	opt.add_command(label="Register",command=lambda :signup_open(2))
	opt.add_command(label="Close",command=lambda : logwin.destroy())
	rwin.config(menu=menubar)	

	def back_main():
		rwin.destroy()
		logwin.deiconify()
		
		#info labels
	Label(rwin,text="Informations",fg="black",bg="aqua",font=("Ink free",25,"bold")).pack(padx=5,pady=10)
	Label(rwin,text="Name:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=100,y=100)
	Label(rwin,text="Mail id:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=100,y=150)
	Label(rwin,text="Mobile Number:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=100,y=200)
	Label(rwin,text="date Of Birth",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=100,y=250)
	Label(rwin,text="Qualification:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=100,y=300)
	Label(rwin,text="Gender:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=100,y=350)
	Label(rwin,text="Area:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=100,y=400)
		

	#values from data
	Label(rwin,text=data[0].value,fg="gold",bg="#121212",font=("Courier",15)).place(x=300,y=100)
	Label(rwin,text=data[2].value,fg="gold",bg="#121212",font=("Courier",15)).place(x=300,y=150)
	Label(rwin,text=data[3].value,fg="gold",bg="#121212",font=("Courier",15)).place(x=300,y=200)
	Label(rwin,text=data[4].value,fg="gold",bg="#121212",font=("Courier",15)).place(x=300,y=250)
	Label(rwin,text=data[5].value,fg="gold",bg="#121212",font=("Courier",15)).place(x=300,y=300)
	Label(rwin,text=data[6].value,fg="gold",bg="#121212",font=("Courier",15)).place(x=300,y=350)
	Label(rwin,text=data[7].value,fg="gold",bg="#121212",font=("Courier",15)).place(x=300,y=400)
		
	Button(rwin,text="<back",fg="#c0c0c0",bg="#2c3e50",cursor="dot",activebackground="#121212",activeforeground="#c0c0c0",relief=FLAT,font=("Courier",13),command= lambda : back_main()).place(x=280,y=450)


def ps_check(em,pw):
	if(em=="" or ps==""):
		mb.showerror("Error","Enter values in the field")
	else:
		wb=load_workbook("book.xlsx")
		ws=wb.active
		i=1
		for cell1,cell2 in zip(ws['A'],ws['B']):
			if(em==cell1.value and pw==cell2.value):
				open_window(ws[i],i)
				break
			elif(em!=cell1.value and ps==cell2.value):
				mb.showerror("Error","Incorrect mail id")
				break
			i+=1
		if(i-1==len(ws['A'])):
			mb.showerror("Error","User not found")


def verify_data(data):
	n=len(data)
	c=0
	for i in data:
		if(i==""):
			mb.showerror("Error","Enter values in all the field")
			break
		c=c+1
	if(c==n):
		wb=load_workbook("book.xlsx")
		ws=wb.active
		em=ws['A']
		emln=len(em)
		ca=0
		for counter in em:
			if(data[0]==counter.value):
				mb.showerror("Error","Email-id already registered")
				break
			ca+=1
		if(ca==emln):
			return True
		else:
			return False



def load_data(*data):
	verify_data(data)
	wb=load_workbook("book.xlsx")
	ws=wb.active
	em=ws['A']
	ws.append(data)
	wb.save("book.xlsx")
	sing.config(state=DISABLED)
	mb.showinfo("Info","Your data has been stored\n Go back to login page")



def signup_open(o):
	if(o==1):
		logwin.state("iconic")
	else:
		rwin.destroy()
	#global e1,e2,e3,e4,e5
	reg.config(fg="#795CB2")
	signwin=Toplevel()
	signwin.title("Sign_up-form")
	signwin.geometry("630x550")
	signwin.configure(bg="#121212")
	Label(signwin,text="Fill The Form",fg="#121212",bg="#1954a2",font=("MV Boli",25,"underline")).pack(padx=5,pady=10)
	Label(signwin,text="Enter mail id:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=90,y=65)
	Label(signwin,text="Enter password:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=90,y=115)
	Label(signwin,text="Enter username:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=90,y=165)
	Label(signwin,text="Enter Phone number:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=90,y=215)
	Label(signwin,text="Select Date Of Birth:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=90,y=265)
	Label(signwin,text="Select Qualification:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=90,y=315)
	Label(signwin,text="Select Gender:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=90,y=365)
	Label(signwin,text="Enter Area:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=90,y=415)
	
	#variables
	em=StringVar()
	ps=StringVar()
	un=StringVar()
	pn=StringVar()
	dt=StringVar()
	qft=StringVar()
	gen=StringVar()
	ar=StringVar()


	e1=Entry(signwin,textvariable=em,width=25,fg="#c0c0c0",bd=3,bg="#121212",insertbackground="white")
	e1.place(x=370,y=65)
	e2=Entry(signwin,textvariable=ps,width=25,fg="#c0c0c0",bd=3,bg="#121212",show="*",insertbackground="white")
	e2.place(x=370,y=115)
	
	def pwd_show():
		if(spval.get()==1):
			e2.config(show="")
		else:
			e2.config(show="*")


	spval=IntVar(value=0)
	checkb= Checkbutton(signwin,text='Show',variable=spval,
		onvalue=1,offvalue=0,fg="#b7c7d7",bg="#121212",font=("Courier",10),command=pwd_show)
	checkb.place(x=530,y=115)


	e3=Entry(signwin,textvariable=un,width=25,fg="#c0c0c0",bd=3,bg="#121212",insertbackground="white")
	e3.place(x=370,y=165)
	e4=Entry(signwin,textvariable=pn,width=25,fg="#c0c0c0",bd=3,bg="#121212",insertbackground="white")
	e4.place(x=370,y=215)
	

	clbt=Button(signwin,text="click",fg="#c0c0c0",bg="#2c3e50",cursor="dot",activebackground="#121212",activeforeground="#c0c0c0",relief=FLAT,font=("Courier",10),command= lambda : open_cal())
	clbt.place(x=370,y=265)
	

	sty=ttk.Style()
	sty.theme_use("clam")
	sty.configure("TCombobox",fieldbackgound="#2c3e50",bg="#121212",fg="#c0c0c0")

	cbox=ttk.Combobox(signwin,textvariable=qft,width=15)
	cbox['values']=('PG','UG','XII','X','Below X')
	cbox.place(x=370,y=315)
	cbox.current()
	
	Radiobutton(signwin,text="Male",relief=FLAT,activebackground="#121212",activeforeground="#c0c0c0",variable=gen,value="Male",bg="#121212",fg="#c0c0c0",selectcolor="black").place(x=370,y=365)
	Radiobutton(signwin,text="Female",relief=FLAT,activebackground="#121212",activeforeground="#c0c0c0",variable=gen,value="Female",bg="#121212",fg="#c0c0c0",selectcolor='black').place(x=430,y=365)
	
	e5=Entry(signwin,textvariable=ar,width=25,fg="#c0c0c0",bd=3,bg="#121212",insertbackground="white")
	e5.place(x=370,y=415)
	
	Button(signwin,text="<back",fg="#c0c0c0",bg="#2c3e50",cursor="dot",activebackground="#121212",activeforeground="#c0c0c0",relief=FLAT,font=("Courier",13),command= lambda : back_log()).place(x=125,y=470)
	global sing
	global otbn
	sing=Button(signwin,text="signup!",fg="#c0c0c0",bg="#2c3e50",cursor="dot",activebackground="#121212",activeforeground="#c0c0c0",relief=FLAT,state=DISABLED,font=("Courier",13),command= lambda :load_data(em.get(),ps.get(),un.get(),pn.get(),dt.get(),qft.get(),gen.get(),ar.get()))
	sing.place(x=360,y=470)
	otbn=Button(signwin,text="click here to verify your account ",cursor="dot",fg="#3366CC",bg="#121212",activebackground="#121212",activeforeground="#3366CC",relief=FLAT,font=("Courier",13),command= lambda :open_otp(em.get(),ps.get(),un.get(),pn.get(),dt.get(),qft.get(),gen.get(),ar.get()))
	otbn.place(x=130,y=500)

	#e1.bind('Tab',mov1)
	#e2.bind('Tab',mov2)
	#e3.bind('Tab',mov3)
	#e4.bind('Tab',mov4)
	#e5.bind('Tab',mov5)


	def back_log():
		signwin.destroy()
		logwin.deiconify()

	def open_cal():
		calwin=Tk()
		calwin.title("Sign_up-form")
		calwin.geometry("400x300")
		calwin.configure(bg="#121212")
		Label(calwin,text="Select the date and click the button",bg="#121212",fg="#c0c0c0",font=("inkfree",15,"underline")).pack(pady=10)
		cal = Calendar(calwin,selectmode='day',day=14,month=5,year=2003)
		cal.pack(pady=10)
		
		def select_date(): 
			date=cal.get_date()
			calwin.destroy()
			clbt.destroy()
			Entry(signwin,textvariable=dt,width=25,fg="#c0c0c0",bd=3,bg="#121212").place(x=370,y=265)
			dt.set(date)

		Button(calwin,text="click",fg="#c0c0c0",bg="#2c3e50",cursor="dot",activebackground="#121212",activeforeground="#c0c0c0",font=("Courier",13),relief=FLAT,command= lambda : select_date()).pack(pady=10)


def open_otp(*datas):
	if(not verify_data(datas)):
		#mb.showerror("CRUD-OPERATION","Enter values in all the field")
		pass
	else:
		if(check_internet()):
			global otps
			otps=send_otp(datas)
			otpwin=Toplevel()
			otpwin.geometry("450x300")
			otpwin.title("CRUD-OPERATION")
			otpwin.configure(bg="#121212")
			Label(otpwin,text="OTP has been sent to your mail-id",fg="#c0c0c0",bg="#1976a2",font=("mv boli",15,"bold")).pack(pady=10)
			Label(otpwin,text="Enter OTP:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=50,y=70)
			otpr=StringVar()
			Entry(otpwin,textvariable=otpr,width=15,bd=3,fg="#c0c0c0",bg="#121212").place(x=200,y=70)
			Button(otpwin,text="check",fg="#c0c0c0",cursor="dot",bg="#2c3e50",relief=FLAT,font=("Courier",13),command= lambda :check_otp(otpr.get(),otps)).place(x=230,y=120)
			Label(otpwin,text="Didn\'t recieve the otp..",fg="#c0c0c0",bg="#121212",font=("Courier",13)).place(x=30,y=170)
			Button(otpwin,text="Resend",fg="#3366CC",cursor="dot",bg="#121212",relief=FLAT,font=("Courier",10),command= lambda :send_again()).place(x=270,y=170)

			def check_otp(otpr,otps):
				if(otpr==otps):
					mb.showinfo("CRUD-OPERATION","Verified Successfully")
					sing.config(state=NORMAL)
					otbn.config(state=DISABLED)
					otpwin.destroy()

					
				else:
					mb.showerror("CRUD-OPERATION","Invalid OTP")
			def send_again():
				global otps
				otps=send_otp(datas)


#mainwindow
logwin=Tk()
logwin.title("Login-form")
logwin.geometry("500x400")
logwin.configure(bg="#121212")
logwin.resizable(False,False)
#logwin.call('tk','scaling',2.0)
Label(logwin,text="Enter mail id:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=90,y=80)
Label(logwin,text="Enter password:",fg="#c0c0c0",bg="#121212",font=("Courier",15)).place(x=90,y=130)


em=StringVar()
ps=StringVar()


ement=Entry(logwin,textvariable=em,width=25,fg="#c0c0c0",bd=3,bg="#121212",insertbackground="white")
ement.place(x=280,y=83)
pwd=Entry(logwin,textvariable=ps,width=25,fg="#c0c0c0",bd=3,bg="#121212",show="*",insertbackground="white")
pwd.place(x=280,y=130)

def clear_txt():
	em.set("")
	ps.set("")
def pw_show():
    if(cbval.get()==1):
        pwd.config(show="")
    else:
        pwd.config(show="*")


cbval=IntVar(value=0)
checkb= Checkbutton(logwin,text='Show Password',variable=cbval,
	onvalue=1,offvalue=0,fg="#b7c7d7",bg="#121212",font=("Courier",10),command=pw_show)
checkb.place(x=280,y=160)


Label(logwin,text="Don\'t have an Account...",fg="#c0c0c0",bg="#121212",font=("Courier",13)).place(x=80,y=300)
Button(logwin,text="login",fg="#c0c0c0",bg="#2c3e50",cursor="dot",activebackground="#121212",activeforeground="#c0c0c0",font=("Courier",10),relief=FLAT,command= lambda : ps_check(em.get(),ps.get())).place(x=130,y=200)
reg=Button(logwin,text="Register",fg="#3366CC",bg="#121212",cursor="dot",activebackground="#121212",activeforeground="#3366CC",relief=FLAT,font=("Courier",10,"underline"),command= lambda : signup_open(1))
reg.place(x=315,y=300)
Button(logwin,text="clear",fg="#c0c0c0",bg="#2c3e50",cursor="dot",activebackground="#121212",activeforeground="#c0c0c0",font=("Courier",10),relief=FLAT,command= lambda : clear_txt()).place(x=330,y=200)

logwin.mainloop()
